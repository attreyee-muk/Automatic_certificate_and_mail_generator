import cv2 as cv
import openpyxl


# provide path of the Excel sheet you want to read data from.
obj = openpyxl.load_workbook('')
sheet = obj.active

row_count = sheet.max_row

# provide path to the file where your certificate template is located
img = cv.imread('')

font = cv.FONT_HERSHEY_TRIPLEX
font_size = 2.5
font_color = (0, 0, 0)
x = 105
y = 500
# provide the output path where the certificates need to be stored.
output_path = './certificates'


for i in range(1, row_count+1):
    name = sheet.cell(i, 6).value
    # provide path to the file where your certificate template is located
    img = cv.imread('')

    font = cv.FONT_HERSHEY_SIMPLEX

    cv.putText(img, name,
               (x, y),
               font,
               font_size,
               font_color, 10)
    certi_path = "./certificates/" + name + '.jpg'
    cv.imwrite(certi_path, img)
