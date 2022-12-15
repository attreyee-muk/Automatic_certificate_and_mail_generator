import smtplib
import ssl
from email.mime.text import MIMEText
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import openpyxl
import imghdr

# provide path of the Excel sheet you want to read data from.
obj = openpyxl.load_workbook('')
sheet = obj.active

row_count = sheet.max_row
subject = " "  # add subject
sender = " "  # add sender email id
password = " "  # add password key generated by gmail

try:
    for i in range(1, row_count):
        # add your the row number and column number in sheet.cell(row no. , column no.)
        receiver = sheet.cell(i + 1, 8).value
        # add your the row number and column number in sheet.cell(row no. , column no.)
        name = sheet.cell(i + 1, 6).value
        message = MIMEMultipart()
        message["From"] = sender
        message["To"] = receiver
        message["Subject"] = subject
        heading = f'Hi {name}!'
        text = """\
            <html>
                <body>
                <p>
                Add your content here
                </p>
                </body>
            </html>
        """
        message.attach(MIMEText(heading, "plain"))
        message.attach(MIMEText(text, "html"))
        text = message.as_string()

        # message.add_attachment(image_data, maintype='image', subtype=image_type, filename=image_name)
        context = ssl.create_default_context()
        try:
            with smtplib.SMTP_SSL('smtp.gmail.com', 465, context=context) as server:
                server.login(sender, password)
                server.sendmail(sender, receiver, text)
            print(name, "Successful")
        except Exception as e:
            print("email not sent")

except Exception as e:
    print("failed")
